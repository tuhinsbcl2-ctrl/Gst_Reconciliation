"""
reconciliation/reports.py — Excel report generation for GST Reconciliation App.

All public functions accept DataFrames (or file paths) and write a formatted
Excel workbook to *output_path*.  Returns the resolved output path as a string.
"""

import os
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

import config
from reconciliation.engine import (
    COL_GSTIN,
    COL_INVOICE_NO,
    COL_PERIOD,
    COL_TAX_RATE,
    COL_TAXABLE_VALUE,
    COL_TOTAL_TAX,
    CAT_MATCHED,
    CAT_MISMATCH,
    reconcile_gstr1_vs_gstr3b,
    reconcile_gstr2b_vs_gstr3b,
    reconcile_gstr2a_vs_gstr2b,
)


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_SUBHEADER_FILL = PatternFill("solid", fgColor="2E75B6")
_SUBHEADER_FONT = Font(bold=True, color="FFFFFF")
_MISMATCH_FILL = PatternFill("solid", fgColor="FFD7D7")
_MATCHED_FILL = PatternFill("solid", fgColor="D7F5D7")
_ALT_ROW_FILL = PatternFill("solid", fgColor="EBF3FB")

_THIN = Side(border_style="thin", color="AAAAAA")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center")


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _apply_header(ws, row: int, headers: list[str]) -> None:
    """Write bold white-on-blue header row."""
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _CELL_BORDER


def _apply_data_row(ws, row: int, values: list, mismatch: bool = False) -> None:
    """Write a data row with alternating fill and optional mismatch highlight."""
    fill = _MISMATCH_FILL if mismatch else (_ALT_ROW_FILL if row % 2 == 0 else None)
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        if fill:
            cell.fill = fill
        cell.alignment = _LEFT
        cell.border = _CELL_BORDER


def _auto_column_width(ws) -> None:
    """Auto-fit column widths based on max content length."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)


def _default_output(name: str, output_path: Optional[str]) -> Path:
    """Return *output_path* as Path, or a default path in the REPORTS_DIR."""
    if output_path:
        return Path(output_path)
    return config.REPORTS_DIR / f"{name}.xlsx"


# ---------------------------------------------------------------------------
# 1. GSTR-1 Detailed Report (Rate-wise & Party-wise)
# ---------------------------------------------------------------------------

def gstr1_detailed_report(
    gstr1_df: pd.DataFrame,
    output_path: Optional[str] = None,
    party_col: str = "PartyName",
) -> str:
    """Generate GSTR-1 Detailed Report with Rate-wise and Party-wise summaries.

    Args:
        gstr1_df:    DataFrame loaded from GSTR-1 Excel.  Expected columns:
                     [Period, GSTIN, PartyName, TaxRate, TaxableValue, TotalTax].
        output_path: Where to save the output workbook.
        party_col:   Column name for party/recipient name.

    Returns:
        Absolute path to the saved Excel file.
    """
    out = _default_output("GSTR1_Detailed_Report", output_path)
    numeric_cols = [COL_TAXABLE_VALUE, COL_TOTAL_TAX]

    df = gstr1_df.copy()
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ---- Rate-wise summary ----
    ws_rate = wb.create_sheet("Rate-wise Summary")
    _apply_header(ws_rate, 1, ["Tax Rate (%)", "Period", "Taxable Value", "Total Tax"])
    row = 2
    if COL_TAX_RATE in df.columns and COL_PERIOD in df.columns:
        rate_summary = (
            df.groupby([COL_TAX_RATE, COL_PERIOD])[[COL_TAXABLE_VALUE, COL_TOTAL_TAX]]
            .sum()
            .reset_index()
            .sort_values([COL_TAX_RATE, COL_PERIOD])
        )
        for _, r in rate_summary.iterrows():
            _apply_data_row(
                ws_rate, row,
                [r[COL_TAX_RATE], r[COL_PERIOD], round(r[COL_TAXABLE_VALUE], 2), round(r[COL_TOTAL_TAX], 2)],
            )
            row += 1
    _auto_column_width(ws_rate)

    # ---- Party-wise summary ----
    ws_party = wb.create_sheet("Party-wise Summary")
    _apply_header(ws_party, 1, [party_col, "GSTIN", "Total Taxable Value", "Total Tax"])
    row = 2
    group_cols = [party_col, COL_GSTIN] if COL_GSTIN in df.columns else [party_col]
    if party_col in df.columns:
        party_summary = (
            df.groupby(group_cols)[[COL_TAXABLE_VALUE, COL_TOTAL_TAX]]
            .sum()
            .reset_index()
            .sort_values(party_col)
        )
        for _, r in party_summary.iterrows():
            vals = [r[party_col], r.get(COL_GSTIN, ""), round(r[COL_TAXABLE_VALUE], 2), round(r[COL_TOTAL_TAX], 2)]
            _apply_data_row(ws_party, row, vals)
            row += 1
    _auto_column_width(ws_party)

    # ---- Monthly summary ----
    ws_monthly = wb.create_sheet("Monthly Summary")
    _apply_header(ws_monthly, 1, ["Period", "Total Taxable Value", "Total Tax"])
    row = 2
    if COL_PERIOD in df.columns:
        monthly = (
            df.groupby(COL_PERIOD)[[COL_TAXABLE_VALUE, COL_TOTAL_TAX]]
            .sum()
            .reset_index()
            .sort_values(COL_PERIOD)
        )
        for _, r in monthly.iterrows():
            _apply_data_row(ws_monthly, row, [r[COL_PERIOD], round(r[COL_TAXABLE_VALUE], 2), round(r[COL_TOTAL_TAX], 2)])
            row += 1
    _auto_column_width(ws_monthly)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)


# ---------------------------------------------------------------------------
# 2. GSTR-3B Detailed Report
# ---------------------------------------------------------------------------

def gstr3b_detailed_report(
    gstr3b_df: pd.DataFrame,
    output_path: Optional[str] = None,
) -> str:
    """Generate GSTR-3B Detailed Report with monthly summary.

    Args:
        gstr3b_df:   DataFrame from GSTR-3B Excel.  Expected columns:
                     [Period, TaxableValue, TotalTax, ITCClaimed].
        output_path: Where to save the output workbook.

    Returns:
        Absolute path to the saved Excel file.
    """
    out = _default_output("GSTR3B_Detailed_Report", output_path)
    df = gstr3b_df.copy()
    numeric_cols = [COL_TAXABLE_VALUE, COL_TOTAL_TAX]
    if "ITCClaimed" in df.columns:
        numeric_cols.append("ITCClaimed")
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

    wb = Workbook()
    ws = wb.active
    ws.title = "GSTR-3B Monthly Summary"

    headers = ["Period", "Taxable Value (Outward)", "Total Tax Liability"]
    if "ITCClaimed" in df.columns:
        headers.append("ITC Claimed")
    _apply_header(ws, 1, headers)

    row = 2
    if COL_PERIOD in df.columns:
        group_cols = [c for c in [COL_TAXABLE_VALUE, COL_TOTAL_TAX, "ITCClaimed"] if c in df.columns]
        monthly = (
            df.groupby(COL_PERIOD)[group_cols]
            .sum()
            .reset_index()
            .sort_values(COL_PERIOD)
        )
        for _, r in monthly.iterrows():
            vals = [r[COL_PERIOD], round(r[COL_TAXABLE_VALUE], 2), round(r[COL_TOTAL_TAX], 2)]
            if "ITCClaimed" in r:
                vals.append(round(r["ITCClaimed"], 2))
            _apply_data_row(ws, row, vals)
            row += 1

    _auto_column_width(ws)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)


# ---------------------------------------------------------------------------
# 3. GSTR-1 vs GSTR-3B Reconciliation Report
# ---------------------------------------------------------------------------

def gstr1_vs_gstr3b_report(
    gstr1_df: pd.DataFrame,
    gstr3b_df: pd.DataFrame,
    output_path: Optional[str] = None,
) -> str:
    """Write GSTR-1 vs GSTR-3B reconciliation to Excel.

    Returns absolute path of saved file.
    """
    out = _default_output("GSTR1_vs_GSTR3B_Reco", output_path)
    result = reconcile_gstr1_vs_gstr3b(gstr1_df, gstr3b_df)

    wb = Workbook()
    ws = wb.active
    ws.title = "GSTR1 vs GSTR3B"

    headers = list(result.columns)
    _apply_header(ws, 1, headers)

    for row_idx, (_, row) in enumerate(result.iterrows(), start=2):
        is_mismatch = row.get("Status", "") == CAT_MISMATCH
        _apply_data_row(ws, row_idx, list(row), mismatch=is_mismatch)

    _auto_column_width(ws)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)


# ---------------------------------------------------------------------------
# 4. GSTR-2B vs GSTR-3B vs GSTR-2A Combined Report
# ---------------------------------------------------------------------------

def gstr2b_vs_gstr3b_vs_gstr2a_report(
    gstr2b_df: pd.DataFrame,
    gstr3b_df: pd.DataFrame,
    gstr2a_df: pd.DataFrame,
    output_path: Optional[str] = None,
) -> str:
    """Three-way comparison: GSTR-2B, GSTR-3B, and GSTR-2A.

    Writes two sheets:
        - '2B vs 3B'  : ITC available vs ITC claimed
        - '2A vs 2B'  : Invoice-level matching

    Returns absolute path of saved file.
    """
    out = _default_output("GSTR2B_vs_GSTR3B_vs_GSTR2A", output_path)
    result_2b_3b = reconcile_gstr2b_vs_gstr3b(gstr2b_df, gstr3b_df)
    result_2a_2b = reconcile_gstr2a_vs_gstr2b(gstr2a_df, gstr2b_df)

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: 2B vs 3B
    ws1 = wb.create_sheet("2B vs 3B")
    _apply_header(ws1, 1, list(result_2b_3b.columns))
    for row_idx, (_, row) in enumerate(result_2b_3b.iterrows(), start=2):
        is_mismatch = row.get("Status", "") not in (CAT_MATCHED, "Under-claimed ITC")
        _apply_data_row(ws1, row_idx, list(row), mismatch=is_mismatch)
    _auto_column_width(ws1)

    # Sheet 2: 2A vs 2B
    ws2 = wb.create_sheet("2A vs 2B")
    _apply_header(ws2, 1, list(result_2a_2b.columns))
    for row_idx, (_, row) in enumerate(result_2a_2b.iterrows(), start=2):
        is_mismatch = row.get("Status", "") != CAT_MATCHED
        _apply_data_row(ws2, row_idx, list(row), mismatch=is_mismatch)
    _auto_column_width(ws2)

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)


# ---------------------------------------------------------------------------
# 5. Financial Year-wise Summary
# ---------------------------------------------------------------------------

def financial_year_summary(
    df: pd.DataFrame,
    financial_year: str = "",
    output_path: Optional[str] = None,
) -> str:
    """Aggregate data across months for a selected financial year (April–March).

    Args:
        df:             DataFrame with at least [Period, TaxableValue, TotalTax].
                        Period should be in 'YYYY-MM' format.
        financial_year: Label like '2023-24'.  If empty, all data is used.
        output_path:    Where to save the workbook.

    Returns:
        Absolute path of saved file.
    """
    out = _default_output(f"FY_Summary_{financial_year or 'All'}", output_path)
    _df = df.copy()
    for col in [COL_TAXABLE_VALUE, COL_TOTAL_TAX]:
        if col in _df.columns:
            _df[col] = pd.to_numeric(_df[col], errors="coerce").fillna(0)

    # Filter to financial year if provided
    if financial_year and COL_PERIOD in _df.columns:
        try:
            start_year, end_year = financial_year.split("-")
            # Apr-Mar spanning two calendar years
            fy_periods: list[str] = []
            for m in range(4, 13):
                fy_periods.append(f"{start_year}-{m:02d}")
            for m in range(1, 4):
                fy_periods.append(f"20{end_year}-{m:02d}")
            _df = _df[_df[COL_PERIOD].isin(fy_periods)]
        except Exception:
            pass  # fall through with unfiltered data

    wb = Workbook()
    ws = wb.active
    ws.title = f"FY Summary {financial_year}"
    _apply_header(ws, 1, ["Period", "Total Taxable Value", "Total Tax"])

    if COL_PERIOD in _df.columns:
        monthly = (
            _df.groupby(COL_PERIOD)[[COL_TAXABLE_VALUE, COL_TOTAL_TAX]]
            .sum()
            .reset_index()
            .sort_values(COL_PERIOD)
        )
        for row_idx, (_, row) in enumerate(monthly.iterrows(), start=2):
            _apply_data_row(ws, row_idx, [row[COL_PERIOD], round(row[COL_TAXABLE_VALUE], 2), round(row[COL_TOTAL_TAX], 2)])

        # Grand total row
        total_row = len(monthly) + 2
        ws.cell(total_row, 1, "Grand Total").font = Font(bold=True)
        ws.cell(total_row, 2, round(monthly[COL_TAXABLE_VALUE].sum(), 2)).font = Font(bold=True)
        ws.cell(total_row, 3, round(monthly[COL_TOTAL_TAX].sum(), 2)).font = Font(bold=True)

    _auto_column_width(ws)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)


# ---------------------------------------------------------------------------
# 6. Monthly Summary
# ---------------------------------------------------------------------------

def monthly_summary(
    df: pd.DataFrame,
    output_path: Optional[str] = None,
) -> str:
    """Month-by-month breakdown of all return data.

    Args:
        df:          DataFrame with [Period, TaxableValue, TotalTax].
        output_path: Where to save the workbook.

    Returns:
        Absolute path of saved file.
    """
    out = _default_output("Monthly_Summary", output_path)
    _df = df.copy()
    for col in [COL_TAXABLE_VALUE, COL_TOTAL_TAX]:
        if col in _df.columns:
            _df[col] = pd.to_numeric(_df[col], errors="coerce").fillna(0)

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Summary"
    _apply_header(ws, 1, ["Month", "Taxable Value", "Total Tax"])

    if COL_PERIOD in _df.columns:
        monthly = (
            _df.groupby(COL_PERIOD)[[COL_TAXABLE_VALUE, COL_TOTAL_TAX]]
            .sum()
            .reset_index()
            .sort_values(COL_PERIOD)
        )
        for row_idx, (_, row) in enumerate(monthly.iterrows(), start=2):
            _apply_data_row(ws, row_idx, [row[COL_PERIOD], round(row[COL_TAXABLE_VALUE], 2), round(row[COL_TOTAL_TAX], 2)])

    _auto_column_width(ws)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return str(out)
